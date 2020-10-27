import openpyxl


class ExcelHandler:

    def __init__(self, file_name):
        """初始化"""
        self.file_name = file_name

    def open_file(self):
        """打开文件"""
        workbook = openpyxl.load_workbook(self.file_name)
        self.workbook = workbook
        return workbook

    def get_sheet(self, name):
        """获取表格"""
        workbook = self.open_file()
        return workbook[name]

    def read_data(self, sheet_name):
        """读取数据. 每一行数据存放到列表或者是字典当中。"""
        sheet = self.get_sheet(sheet_name)

        # 所有的行
        rows = list(sheet.rows)

        data = []
        # 获取标题， 作为每一行数据字典的 key
        headers = []
        for title in rows[0]:
            headers.append(title.value)

        # 添加数据
        for row in rows[1:]:
            row_data = {}

            for idx, cell in enumerate(row):
                row_data[headers[idx]] = cell.value

            data.append(row_data)

        self.close()
        return data

    def write(self, sheet_name, row_num, column_num, data):
        """写入单元格数据到第row，column单元格"""
        sheet = self.get_sheet(sheet_name)
        sheet.cell(row_num, column_num).value = data
        self.save()
        self.close()

    def save(self):
        self.workbook.save(self.file_name)

    def close(self):
        self.workbook.close()


if __name__ == '__main__':
    pass
    xls = ExcelHandler("./../test.xlsx")
    sheet = xls.get_sheet("register")
    print(xls.read_data('register'))
    print(sheet.cell(1,2).value)